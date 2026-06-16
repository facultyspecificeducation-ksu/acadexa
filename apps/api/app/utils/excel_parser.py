"""
================================================================================
Excel Parser Module for Acadexa Academic Advisory System
================================================================================

This module provides comprehensive Excel file parsing functionality for student
academic records. It extracts raw student data, course information, and grades
from Excel files, then calculates GPAs and academic statistics.

Key Features:
-------------
- Parses student information (ID, name, study level, cumulative percentage)
- Extracts semester data with automatic column offset detection
- Reads course grades in English format (A, B+, C, etc.)
- Maps grades to grade points from database grade_scale table
- Handles special grades from grade_scale (W, AU, S, MW, FW, EX, IC, TC)
- Calculates semester GPA, cumulative GPA, attempted vs completed hours
- Matches students to correct curriculum based on department + enrollment year
- Matches courses to curriculum_courses for prerequisite tracking
- Saves parsed data to Supabase with upsert operations

Database Integration (v3.0):
----------------------------
- Uses grade_scale table for grade points (not hardcoded grade_map)
- Resolves curriculum_id for each student (department_id + enrollment_year)
- Matches each course to curriculum_course_id for prerequisite validation
- Stores level_semester_raw, term, and level in student_semesters
- Uses correct column names: grade_score (not score), curriculum_course_id, etc.

Author: Acadexa Team
Version: 3.1.0
Last Modified: 2026-06-16
Made with ❤️ for Faculty of Specific Education, Kafr El-Sheikh University

================================================================================
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional
from datetime import datetime
import logging

logger = logging.getLogger("acadexa.parser")

# =============================================================================
# Helper Functions
# =============================================================================

def clean(value):
    """Clean and normalize cell values - removes whitespace and handles None."""
    if value is None:
        return ""
    return str(value).strip()

def get_cell(ws, col_letter, row_idx):
    """Safely get cell value from worksheet by column letter and row index."""
    try:
        return ws[f"{col_letter}{row_idx}"].value
    except Exception:
        return None

def parse_float(value) -> float | None:
    """
    Safely convert any value to float.

    Returns None (not 0.0) when the value is missing or blank so callers can
    distinguish "no data provided" from an actual zero.  Callers that need a
    fallback should use:  parse_float(x) or 0.0
    """
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).strip())
    except Exception:
        return None

def parse_int(value) -> int | None:
    """
    Safely convert any value to integer.

    Returns None (not 0) when the value is missing or blank so callers can
    distinguish "no credit hours recorded" from an actual zero-hour entry.
    Callers that need a fallback should use:  parse_int(x) or 0
    """
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value)))
    except Exception:
        return None

def parse_percentage(value) -> float:
    """
    Safely parse percentage string like '72.5%' to float.
    Returns None if parsing fails.
    """
    if not value:
        return None
    try:
        return float(str(value).replace("%", "").strip())
    except Exception:
        return None

def normalize_passed(value) -> bool:
    """
    Normalize pass/fail values from Excel to boolean.

    Accepts various formats:
      Arabic : 'نعم', 'ناجح'
      Latin  : 'yes', 'pass', 'true', '1'
      Symbols: '✔', '✓', '√'   ← '√' is the actual mark used in these sheets

    Anything not in the list is treated as False (not passed / not yet graded).
    """
    if not value:
        return False

    normalised = clean(str(value)).lower()
    passed_keywords = ["نعم", "ناجح", "yes", "pass", "true", "1", "✔", "✓", "√"]
    return normalised in passed_keywords

# =============================================================================
# Grade Scale Functions (Database-Driven)
# =============================================================================

def get_grade_info_from_db(supabase, grade_letter: str) -> Optional[Dict]:
    """
    Fetch grade information (points, affects_gpa, is_passing) from grade_scale table.
    
    Args:
        supabase: Supabase client instance
        grade_letter: The grade letter (e.g., 'A', 'B+', 'W')
    
    Returns:
        Dictionary with keys: points, affects_gpa, is_passing, name_ar
        Returns None if grade not found or supabase not available
    """
    if not supabase or not grade_letter:
        return None
    
    try:
        result = supabase.table("grade_scale") \
            .select("points", "affects_gpa", "is_passing", "name_ar") \
            .eq("grade_letter", grade_letter.upper()) \
            .execute()
        
        if result.data:
            return result.data[0]
    except Exception as e:
        logger.warning(f"Could not fetch grade info for {grade_letter}: {e}")
    
    return None

def get_grade_points_from_db(supabase, grade_letter: str) -> float:
    """
    Get grade points from grade_scale table in database.
    
    Args:
        supabase: Supabase client instance
        grade_letter: The grade letter (e.g., 'A', 'B+', 'W')
    
    Returns:
        Grade points as float (0.0 to 4.0), or 0.0 if not found
    """
    grade_info = get_grade_info_from_db(supabase, grade_letter)
    if grade_info:
        return float(grade_info.get("points", 0.0))
    return 0.0

def is_special_grade_from_db(supabase, grade_letter: str) -> bool:
    """
    Check if a grade is a special grade (does not affect GPA) using grade_scale.
    
    Args:
        supabase: Supabase client instance
        grade_letter: The grade letter to check
    
    Returns:
        True if affects_gpa is False, False otherwise
    """
    grade_info = get_grade_info_from_db(supabase, grade_letter)
    if grade_info:
        return not grade_info.get("affects_gpa", True)
    return False

# =============================================================================
# Curriculum Matching Functions
# =============================================================================

def resolve_curriculum_id(supabase, department_id: str, enrollment_year: int) -> Optional[str]:
    """
    Determine the correct curriculum for a student based on department and enrollment year.
    
    Rule: Student follows the curriculum with regulation_year <= enrollment_year,
    choosing the most recent one (largest regulation_year).
    
    Args:
        supabase: Supabase client instance
        department_id: The department UUID
        enrollment_year: Year the student enrolled (e.g., 2022)
    
    Returns:
        Curriculum UUID if found, None otherwise
    """
    if not supabase or not department_id or not enrollment_year:
        return None
    
    try:
        # Find the most recent curriculum with regulation_year <= enrollment_year
        result = supabase.table("curricula") \
            .select("id") \
            .eq("department_id", department_id) \
            .lte("regulation_year", enrollment_year) \
            .order("regulation_year", desc=True) \
            .limit(1) \
            .execute()
        
        if result.data:
            logger.info(f"Resolved curriculum_id {result.data[0]['id']} for department {department_id}, enrollment {enrollment_year}")
            return result.data[0]["id"]
        else:
            logger.warning(f"No curriculum found for department {department_id} with regulation_year <= {enrollment_year}")
    except Exception as e:
        logger.error(f"Error resolving curriculum_id: {e}")
    
    return None

def match_course_to_curriculum(
    supabase, 
    curriculum_id: str, 
    course_code: str, 
    course_name: str
) -> Optional[str]:
    """
    Match a student's course to a curriculum_course record.
    
    Matching strategy:
    1. First try to match by course_code (preferred)
    2. If no match, try to match by course_name (for community-issues course, etc.)
    
    Args:
        supabase: Supabase client instance
        curriculum_id: The curriculum UUID
        course_code: Course code from transcript (may be empty)
        course_name: Course name from transcript
    
    Returns:
        curriculum_course_id if found, None otherwise
    """
    if not supabase or not curriculum_id:
        return None
    
    # Strategy 1: Match by course_code (if available)
    if course_code and course_code.strip():
        try:
            result = supabase.table("curriculum_courses") \
                .select("id") \
                .eq("curriculum_id", curriculum_id) \
                .eq("course_code", course_code.strip()) \
                .execute()
            
            if result.data:
                logger.debug(f"Matched course {course_code} to curriculum_course_id {result.data[0]['id']}")
                return result.data[0]["id"]
        except Exception as e:
            logger.warning(f"Error matching course by code {course_code}: {e}")
    
    # Strategy 2: Match by course_name (fallback)
    if course_name and course_name.strip():
        try:
            result = supabase.table("curriculum_courses") \
                .select("id") \
                .eq("curriculum_id", curriculum_id) \
                .eq("name_ar", course_name.strip()) \
                .execute()
            
            if result.data:
                logger.debug(f"Matched course {course_name} to curriculum_course_id {result.data[0]['id']} (by name)")
                return result.data[0]["id"]
        except Exception as e:
            logger.warning(f"Error matching course by name {course_name}: {e}")
    
    # No match found
    logger.debug(f"No curriculum match for course: code={course_code}, name={course_name}")
    return None

# =============================================================================
# Semester Parsing Helper Functions
# =============================================================================

def parse_term_from_arabic(level_semester_raw: str) -> Optional[str]:
    """
    Parse term_enum value from Arabic semester text.
    
    Args:
        level_semester_raw: e.g., "المستوى الثالث/الفصل الدراسى الأول" or "الفصل الصيفي"
    
    Returns:
        'fall', 'spring', 'summer', or None if cannot parse
    """
    if not level_semester_raw:
        return None
    
    text = level_semester_raw.lower()
    
    if "الصيفي" in text or "صيفي" in text:
        return "summer"
    elif "الثاني" in text or "ثاني" in text:
        return "spring"
    elif "الأول" in text or "أول" in text:
        return "fall"
    
    return None

def parse_level_from_arabic(level_semester_raw: str) -> Optional[int]:
    """
    Parse numeric study level from Arabic semester text.
    
    Args:
        level_semester_raw: e.g., "المستوى الثالث/الفصل الدراسى الأول"
    
    Returns:
        1, 2, 3, 4, or None if cannot parse
    """
    if not level_semester_raw:
        return None
    
    text = level_semester_raw.lower()
    
    if "الأول" in text and "المستوى" in text:
        return 1
    elif "الثاني" in text and "المستوى" in text:
        return 2
    elif "الثالث" in text and "المستوى" in text:
        return 3
    elif "الرابع" in text and "المستوى" in text:
        return 4
    
    return None

# =============================================================================
# Excel Parsing Functions - Offset Detection
# =============================================================================

def find_first_semester_row(ws):
    """
    Find the first row containing semester header information.
    Searches for the Arabic text 'القسم/الشعبة' (Department/Section).
    
    Returns:
        Tuple of (row_index, col_index) or (None, None) if not found
    """
    for row in range(1, min(100, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 30)):
            val = clean(ws.cell(row, col).value)
            if 'القسم/الشعبة' in val:
                return row, col
    return None, None

def detect_offset(ws):
    """
    Detect the offset between the current sheet and the reference sheet layout.
    
    The reference sheet has 'القسم/الشعبة' at row 40, column 11 (K).
    This function calculates the difference to adjust cell positions.
    
    Returns:
        Tuple of (row_offset, col_offset)
    """
    ref_row, ref_col = find_first_semester_row(ws)
    if ref_row and ref_col:
        row_offset = ref_row - 40
        col_offset = ref_col - 11
        return row_offset, col_offset
    return 0, 0

def get_cell_with_offset(ws, base_col_letter, base_row, row_offset, col_offset):
    """
    Get cell value after applying layout offset.

    Args:
        ws: Worksheet object
        base_col_letter: Reference column letter (e.g., 'K')
        base_row: Reference row number
        row_offset: Number of rows to shift
        col_offset: Number of columns to shift

    Returns:
        Cell value at the adjusted position, or None if the resulting
        column index would be out of range (≤ 0).
    """
    base_col_num = openpyxl.utils.column_index_from_string(base_col_letter)
    new_col_num  = base_col_num + col_offset

    # Guard: openpyxl raises ValueError for column index ≤ 0
    if new_col_num < 1:
        logger.warning(
            f"col_offset={col_offset} pushed '{base_col_letter}' to invalid "
            f"column {new_col_num}. Clamping to column 1."
        )
        new_col_num = 1

    new_col_letter = openpyxl.utils.get_column_letter(new_col_num)
    new_row = base_row + row_offset
    return get_cell(ws, new_col_letter, new_row)

# =============================================================================
# Excel Parsing Functions - Student and Course Data
# =============================================================================

def parse_student_info(ws):
    """
    Extract student information from the top rows of the worksheet.

    Looks for specific Arabic labels in first 30 rows:
    - 'كود الطالب :'   → Student ID
    - 'أسم الطالب :'   → Student Name
    - 'مستوى الدراسة :' → Study Level
    - 'النسبة(بحساب النقاط) :' → Cumulative Percentage

    Returns:
        Dict containing student information
    """
    student = {
        "id": "",
        "name": "",
        "study_level": "",
        "cumulative_percentage": "",
    }
    
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            v = clean(cell.value)
            if v.startswith("كود الطالب :"):
                student["id"] = v.replace("كود الطالب :", "").strip()
            elif v.startswith("أسم الطالب :"):
                student["name"] = v.replace("أسم الطالب :", "").strip()
            elif v.startswith("مستوى الدراسة :"):
                student["study_level"] = v.replace("مستوى الدراسة :", "").strip()
            elif v.startswith("النسبة(بحساب النقاط) :"):
                student["cumulative_percentage"] = v.replace("النسبة(بحساب النقاط) :", "").strip()
    
    return student

def parse_semesters_with_offset(ws):
    """
    Parse all semesters and courses from the worksheet using offset detection.
    
    This function:
    1. Detects layout offset to handle varying Excel sheet formats
    2. Identifies semester headers (department, level/semester, academic year)
    3. Extracts course data (code, name, grade, score, hours, pass/fail)
    
    Column mapping for reference layout:
    - K: Department/Section
    - AH: Level/Semester  
    - BO: Academic Year
    - CD: Course sequence number
    - BC: Course name
    - BU: Course code
    - O: Grade letter (A, B+, C, etc.)
    - Q: Score/percentage
    - AL: Credit hours
    - I: Pass/Fail status
    
    Returns:
        List of semester dictionaries, each containing a list of courses
    """
    row_offset, col_offset = detect_offset(ws)
    
    semesters = []
    current_semester = None
    current_courses = []
    
    # Column mappings (reference positions)
    COLS = {
        'K': 'K', 'AH': 'AH', 'BO': 'BO',
        'I': 'I', 'Y': 'Y', 'M': 'M', 'X': 'X',
        'L': 'L', 'W': 'W', 'AN': 'AN',
        'CD': 'CD', 'BC': 'BC', 'BU': 'BU',
        'AL': 'AL', 'Z': 'Z', 'AC': 'AC',
        'AS': 'AS', 'AU': 'AU', 'O': 'O', 'Q': 'Q'
    }
    
    start_row = max(1, 38 + row_offset)
    
    for row_idx in range(start_row, ws.max_row + 1):
        cell_k = get_cell_with_offset(ws, COLS['K'], row_idx, row_offset, col_offset)
        cell_ah = get_cell_with_offset(ws, COLS['AH'], row_idx, row_offset, col_offset)
        cell_bo = get_cell_with_offset(ws, COLS['BO'], row_idx, row_offset, col_offset)
        
        cell_k_clean = clean(cell_k)
        
        # Detect semester header row
        if 'القسم/الشعبة :' in cell_k_clean:
            # Save previous semester if exists
            if current_semester is not None:
                current_semester['courses'] = current_courses
                semesters.append(current_semester)
            
            # Start new semester
            current_semester = {
                "department": cell_k_clean.replace('القسم/الشعبة :', '').strip(),
                "level_semester": clean(cell_ah).replace('المستوى/الفصل :', '').strip(),
                "academic_year": clean(cell_bo).replace('العام الأكاديمي   :', '').strip(),
                "courses": []
            }
            current_courses = []
            continue
        
        if current_semester is None:
            continue
        
        # Parse course data
        cell_cd = clean(get_cell_with_offset(ws, COLS['CD'], row_idx, row_offset, col_offset))
        cell_bc = clean(get_cell_with_offset(ws, COLS['BC'], row_idx, row_offset, col_offset))
        
        # Skip summary/total rows
        if cell_cd == 'م' or cell_bc == 'المجمـــوع' or cell_bc == 'المجموع':
            continue
        
        # Valid course row: sequence number is a positive integer and course name exists.
        try:
            seq_int = int(float(cell_cd))
            is_valid_seq = seq_int > 0
        except (ValueError, TypeError):
            is_valid_seq = False

        if is_valid_seq and cell_bc:
            raw_passed = get_cell_with_offset(ws, COLS['I'], row_idx, row_offset, col_offset)
            raw_grade = get_cell_with_offset(ws, COLS['O'], row_idx, row_offset, col_offset)
            raw_grade_clean = clean(raw_grade)
            
            course = {
                "seq": cell_cd,
                "course_code": clean(get_cell_with_offset(ws, COLS['BU'], row_idx, row_offset, col_offset)),
                "course_name": cell_bc,
                "grade_letter_raw": raw_grade_clean,  # Original grade as it appears
                "grade_letter": raw_grade_clean.upper() if raw_grade_clean else "",  # Normalized to uppercase
                "score": clean(get_cell_with_offset(ws, COLS['Q'], row_idx, row_offset, col_offset)),
                "hours": clean(get_cell_with_offset(ws, COLS['AL'], row_idx, row_offset, col_offset)),
                "passed_raw": clean(raw_passed),  # Original pass/fail text
                "passed": normalize_passed(raw_passed),  # Normalized boolean
            }
            current_courses.append(course)
    
    # Save the last semester
    if current_semester is not None:
        current_semester['courses'] = current_courses
        semesters.append(current_semester)
    
    return semesters

# =============================================================================
# GPA and Statistics Calculation Functions (Using Database Grade Scale)
# =============================================================================

def calculate_semester_stats_with_db(
    courses: List[Dict], 
    supabase, 
    special_grades_cache: Dict = None
) -> Dict[str, Any]:
    """
    Calculate statistics for a single semester using grade_scale from database.
    
    Statistics calculated:
    - GPA: Grade Point Average for the semester
    - attempted_hours: Total credit hours attempted (includes failed courses)
    - completed_hours: Total credit hours successfully completed (passed only)
    - passed_courses: Number of courses passed
    - failed_courses: Number of courses failed
    - total_courses: Total number of courses taken
    - quality_points: Sum of (hours × grade_points)
    - special_cases: List of special grade courses (W, AU, etc.)
    
    IMPORTANT: Pass/Fail is determined by the 'passed' field from Excel (column I),
    NOT by grade points. This ensures consistency with university policies.
    """
    total_quality_points = 0.0
    attempted_hours = 0
    completed_hours = 0
    passed_count = 0
    failed_count = 0
    total_courses = 0
    special_cases = []
    
    for course in courses:
        hours = parse_float(course.get("hours")) or 0.0
        if hours <= 0:
            continue

        grade_letter = course.get("grade_letter", "F")
        grade_points = get_grade_points_from_db(supabase, grade_letter)
        passed = course.get("passed", False)  # From Excel column I — authoritative source
        
        # Check if it's a special grade (does not affect GPA)
        is_special = is_special_grade_from_db(supabase, grade_letter)
        
        # Track special grade courses
        if is_special:
            grade_info = get_grade_info_from_db(supabase, grade_letter)
            special_cases.append({
                "course_name": course.get("course_name", ""),
                "grade_letter": grade_letter,
                "name_ar": grade_info.get("name_ar", "") if grade_info else ""
            })
            total_courses += 1
            continue

        # Regular course — calculate GPA contributions
        total_quality_points += hours * grade_points
        attempted_hours += hours
        total_courses += 1

        # Pass/Fail determination uses Excel data, not grade points
        if passed:
            completed_hours += hours
            passed_count += 1
        else:
            failed_count += 1
    
    if attempted_hours == 0:
        return {
            "gpa": 0.0,
            "attempted_hours": attempted_hours,
            "completed_hours": completed_hours,
            "passed_courses": passed_count,
            "failed_courses": failed_count,
            "total_courses": total_courses,
            "quality_points": 0.0,
            "special_cases": special_cases
        }
    
    return {
        "gpa": round(total_quality_points / attempted_hours, 2),
        "attempted_hours": attempted_hours,
        "completed_hours": completed_hours,
        "passed_courses": passed_count,
        "failed_courses": failed_count,
        "total_courses": total_courses,
        "quality_points": round(total_quality_points, 2),
        "special_cases": special_cases
    }

def calculate_cumulative_stats_with_db(
    semesters: List[Dict], 
    supabase
) -> Dict[str, Any]:
    """
    Calculate cumulative statistics across all semesters using database grade scale.

    For courses that were retaken (identified by resolve_repeated_courses),
    only the latest attempt contributes to the GPA, attempted hours, and
    pass/fail counts.

    IMPORTANT: Pass/Fail determination uses the 'passed' field from Excel,
    NOT grade points. This ensures consistency with university policies.
    """
    total_quality_points = 0.0
    attempted_hours = 0
    completed_hours = 0
    total_passed = 0
    total_failed = 0
    total_courses = 0
    all_special_cases = []
    
    for semester in semesters:
        for course in semester.get("courses", []):
            hours = parse_float(course.get("hours")) or 0.0
            if hours <= 0:
                continue

            grade_letter = course.get("grade_letter", "F")
            grade_points = get_grade_points_from_db(supabase, grade_letter)
            passed = course.get("passed", False)  # From Excel column I — authoritative

            # Always count every course in total_courses for display
            total_courses += 1

            # Track special grade courses (do not affect GPA or hour counts)
            is_special = is_special_grade_from_db(supabase, grade_letter)
            if is_special:
                grade_info = get_grade_info_from_db(supabase, grade_letter)
                all_special_cases.append({
                    "semester": semester.get("academic_year", ""),
                    "course_name": course.get("course_name", ""),
                    "grade_letter": grade_letter,
                    "name_ar": grade_info.get("name_ar", "") if grade_info else ""
                })
                continue

            # For retaken courses, only the latest attempt counts toward GPA/hours
            if not course.get("is_latest_attempt", True):
                continue

            # Regular course — contribute to GPA and hour totals
            total_quality_points += hours * grade_points
            attempted_hours += hours

            if passed:
                completed_hours += hours
                total_passed += 1
            else:
                total_failed += 1
    
    if attempted_hours == 0:
        return {
            "cumulative_gpa": 0.0,
            "attempted_hours": 0,
            "completed_hours": 0,
            "total_passed_courses": 0,
            "total_failed_courses": 0,
            "total_courses": total_courses,
            "completion_rate": 0.0,
            "special_cases": all_special_cases
        }
    
    completion_rate = round((completed_hours / attempted_hours) * 100, 2)
    
    return {
        "cumulative_gpa": round(total_quality_points / attempted_hours, 2),
        "attempted_hours": attempted_hours,
        "completed_hours": completed_hours,
        "total_passed_courses": total_passed,
        "total_failed_courses": total_failed,
        "total_courses": total_courses,
        "completion_rate": completion_rate,
        "special_cases": all_special_cases
    }

# =============================================================================
# Helper Functions for Student Level Calculations
# =============================================================================

def extract_study_level(level_str: str) -> int:
    """Extract numeric study level from Arabic text (e.g., 'الثالث' -> 3)."""
    if not level_str:
        return None
    
    level_map = {"الأول": 1, "الثاني": 2, "الثالث": 3, "الرابع": 4}
    for key, value in level_map.items():
        if key in level_str:
            return value
    return None

def extract_enrollment_year(semesters: List[Dict]) -> int:
    """Extract enrollment year from the first semester's academic year string."""
    if not semesters:
        return None
    
    first_semester = semesters[0]
    academic_year = first_semester.get("academic_year", "")
    if "-" in academic_year:
        try:
            return int(academic_year.split("-")[0])
        except Exception:
            return None
    return None

# =============================================================================
# Repeated-Course Resolution (latest attempt wins for GPA)
# =============================================================================

def _semester_sort_key(sem_idx: int, semester: Dict) -> tuple:
    """
    Build a sort key that reflects true chronological order.

    The academic_year field looks like '2024-2025' and level_semester looks
    like 'المستوى الثالث/الفصل الدراسى الأول' or 'الفصل الصيفي'.

    Sort priority (ascending = oldest first):
      1. Start year extracted from academic_year  (e.g. 2024)
      2. Semester type: Fall=1, Spring=2, Summer=3
         (Fall is الأول, Spring is الثاني, Summer is الصيفي)
      3. Original sheet index as a final tiebreaker

    Returns a tuple that can be compared with < / >.
    """
    academic_year = semester.get("academic_year", "")
    level_semester = semester.get("level_semester", "")

    # Extract the start year from '2024-2025' → 2024
    try:
        start_year = int(academic_year.split("-")[0].strip())
    except (ValueError, IndexError):
        start_year = 0

    # Map semester type to an ordering integer
    sem_lower = level_semester.lower()
    if "الصيفي" in sem_lower or "صيفي" in sem_lower:
        sem_order = 3          # Summer comes after Spring in the same year
    elif "الثاني" in sem_lower or "ثاني" in sem_lower:
        sem_order = 2          # Spring
    else:
        sem_order = 1          # Fall (الأول) or unknown → treat as earliest

    return (start_year, sem_order, sem_idx)

def resolve_repeated_courses(semesters: List[Dict]) -> List[Dict]:
    """
    Identify courses that appear in more than one semester (retakes / summer
    sessions) and mark which attempt should count toward the cumulative GPA.

    Policy (confirmed with university):
    - Only the LATEST attempt's grade and hours count toward the GPA.
    - All attempts are preserved in the raw semester data so academic history
      is never lost.
    - A course is identified by its course_code.

    Chronological order is determined by academic_year + semester type, NOT
    by the sheet index, so an out-of-order workbook still produces correct
    results.  See _semester_sort_key() for the ranking logic.

    This function mutates semesters in place by adding two keys per course:
      - "is_latest_attempt" : bool  → True if this attempt counts for GPA
      - "attempt_number"    : int   → 1 for first attempt, 2 for retake …

    Example:
        2024-2025 Fall  : MATH101 → F   (is_latest_attempt=False, attempt_number=1)
        2024-2025 Summer: MATH101 → C   (is_latest_attempt=True,  attempt_number=2)
        Cumulative GPA uses only the C grade.
    """
    # Build a chronological rank for each semester so we can sort attempts.
    sem_ranks: Dict[int, tuple] = {
        idx: _semester_sort_key(idx, sem)
        for idx, sem in enumerate(semesters)
    }

    # Track: course_code → list of (sort_key, original_index, course_dict)
    appearances: Dict[str, List[tuple]] = {}

    for sem_idx, semester in enumerate(semesters):
        rank = sem_ranks[sem_idx]
        for course in semester.get("courses", []):
            code = course.get("course_code", "").strip()
            if not code:
                # Can't match without a code — always count it
                course["is_latest_attempt"] = True
                course["attempt_number"] = 1
                continue
            appearances.setdefault(code, []).append((rank, sem_idx, course))

    for code, attempts in appearances.items():
        # Sort by the chronological rank (year → semester type → original index)
        attempts_sorted = sorted(attempts, key=lambda x: x[0])

        for attempt_num, (_, _, course) in enumerate(attempts_sorted, start=1):
            is_latest = (attempt_num == len(attempts_sorted))
            course["is_latest_attempt"] = is_latest
            course["attempt_number"] = attempt_num

            if not is_latest:
                logger.debug(
                    f"Course {code} attempt {attempt_num} "
                    f"(year/sem rank {attempts_sorted[attempt_num-1][0]}) "
                    f"marked non-counting — superseded by a later attempt."
                )

    return semesters

# =============================================================================
# Main ExcelParser Class (UPDATED for Database Schema v3)
# =============================================================================

class ExcelParser:
    """
    Main parser class for extracting and processing student academic data from Excel files.

    This class handles:
    - Reading Excel workbooks with multiple sheets
    - Parsing student information and course data
    - Calculating GPAs and academic statistics using database grade_scale
    - Matching students to curricula based on department + enrollment year
    - Matching courses to curriculum_courses for prerequisite validation
    - Saving processed data to Supabase with upsert operations

    Usage Example:
        parser = ExcelParser(Path("student_records.xlsx"), department_code="CS")
        students, errors = parser.parse_all_students()
        stats = parser.get_stats()

    Attributes:
        file_path: Path to the Excel file
        department_code: Department code for filtering (optional)
        stats: Parsing statistics counter
        errors: List of errors encountered during parsing
    """

    def __init__(self, file_path: Path, department_code: str = ""):
        """
        Initialize the ExcelParser.

        Args:
            file_path: Path to the Excel file to parse
            department_code: Department code for the students (optional)
        """
        self.file_path = file_path
        self.department_code = department_code
        self.stats = {"students": 0, "courses": 0, "semesters": 0}
        self.errors = []
        self._supabase = None
        self._department_id = None
        self._curriculum_cache = {}  # Cache for curriculum lookups by department+year
        self._course_match_cache = {}  # Cache for course matching
    
    def _get_supabase(self):
        """
        Lazy load Supabase client.
        
        Uses the project's Supabase utilities via app.core.supabase.
        Returns None if not available (for testing).
        """
        if self._supabase is None:
            try:
                from app.core.supabase import get_service_role_client
                self._supabase = get_service_role_client()
                logger.info("Supabase service role client initialized for parser")
            except ImportError as e:
                logger.warning(f"Supabase client not available: {e}")
                self._supabase = None
            except Exception as e:
                logger.warning(f"Failed to initialize Supabase client: {e}")
                self._supabase = None
        return self._supabase
    
    def _get_department_id(self):
        """Get department ID from department code with caching."""
        if self._department_id is not None:
            return self._department_id
        
        if not self.department_code:
            return None
        
        try:
            supabase = self._get_supabase()
            if supabase:
                result = supabase.table("departments").select("id").eq("code", self.department_code).execute()
                if result.data:
                    self._department_id = result.data[0]["id"]
                    logger.info(f"Found department ID {self._department_id} for code {self.department_code}")
                    return self._department_id
        except Exception as e:
            logger.warning(f"Could not find department for code {self.department_code}: {e}")
        
        return None
    
    def _resolve_curriculum_with_cache(self, department_id: str, enrollment_year: int) -> Optional[str]:
        """Resolve curriculum_id with caching to avoid repeated DB calls."""
        cache_key = f"{department_id}:{enrollment_year}"
        if cache_key in self._curriculum_cache:
            return self._curriculum_cache[cache_key]
        
        curriculum_id = resolve_curriculum_id(self._get_supabase(), department_id, enrollment_year)
        self._curriculum_cache[cache_key] = curriculum_id
        return curriculum_id
    
    def _match_course_with_cache(
        self, 
        curriculum_id: str, 
        course_code: str, 
        course_name: str
    ) -> Optional[str]:
        """Match course with caching to avoid repeated DB calls."""
        cache_key = f"{curriculum_id}:{course_code}:{course_name}"
        if cache_key in self._course_match_cache:
            return self._course_match_cache[cache_key]
        
        course_id = match_course_to_curriculum(
            self._get_supabase(), 
            curriculum_id, 
            course_code, 
            course_name
        )
        self._course_match_cache[cache_key] = course_id
        return course_id
    
    def parse_all_students(self) -> Tuple[List[Dict], List[Dict]]:
        """
        Parse all students from all sheets in the Excel file.
        
        This is the main entry point for parsing. It:
        1. Opens the Excel workbook
        2. Iterates through each sheet
        3. Extracts student info and course data
        4. Calculates GPA and statistics using database grade_scale
        5. Matches students to curricula and courses to curriculum_courses
        6. Saves to Supabase (if available)
        7. Returns parsed data and any errors encountered
        
        Returns:
            Tuple of (students_data_list, errors_list)
        """
        students = []
        errors = []
        
        supabase = self._get_supabase()
        
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            logger.info(f"Opened workbook with {len(wb.sheetnames)} sheets")
            
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    
                    # Parse raw data from sheet
                    student_info = parse_student_info(ws)
                    semesters = parse_semesters_with_offset(ws)

                    # Mark which attempt of each repeated course counts for GPA
                    resolve_repeated_courses(semesters)
                    
                    # Calculate cumulative statistics using database grade_scale
                    cumulative_stats = calculate_cumulative_stats_with_db(semesters, supabase)
                    
                    # Calculate per-semester statistics using database grade_scale
                    for semester in semesters:
                        semester["calculated"] = calculate_semester_stats_with_db(
                            semester.get("courses", []), 
                            supabase
                        )
                    
                    # Additional metadata
                    study_level_num = extract_study_level(student_info.get("study_level", ""))
                    enrollment_year = extract_enrollment_year(semesters)

                    # Build complete student data structure
                    student_data = {
                        "raw": {
                            "student": student_info,
                            "semesters": semesters,
                        },
                        "calculated": cumulative_stats,
                        "metadata": {
                            "sheet_name": sheet_name,
                            "parsed_at": datetime.utcnow().isoformat(),
                            "study_level": study_level_num,
                            "enrollment_year": enrollment_year,
                        }
                    }
                    
                    # Validate student has an ID
                    if student_info.get("id"):
                        students.append(student_data)
                        self.stats["students"] += 1
                        self.stats["semesters"] += len(semesters)
                        for sem in semesters:
                            self.stats["courses"] += len(sem.get("courses", []))
                        
                        # Save to Supabase if available
                        if supabase:
                            self._save_to_supabase(student_data)
                    else:
                        errors.append({
                            "sheet": sheet_name,
                            "error": "No valid student ID found"
                        })
                        
                except Exception as e:
                    error_msg = str(e)
                    logger.error(f"Error parsing sheet {sheet_name}: {error_msg}")
                    errors.append({
                        "sheet": sheet_name,
                        "error": error_msg
                    })
                    continue
            
            wb.close()

            # Merge any DB-save errors
            if self.errors:
                errors.extend(self.errors)
                self.errors = []

            logger.info(f"Parsing complete: {self.stats}")
            
        except Exception as e:
            logger.error(f"Failed to open workbook: {e}")
            errors.append({
                "sheet": "all",
                "error": f"Failed to open workbook: {str(e)}"
            })
        
        return students, errors
    
    def _save_to_supabase(self, student_data: Dict):
        """
        Save parsed student data to Supabase with upsert operations.
        
        This method is fully compatible with the database schema v3:
        1. Resolves curriculum_id for the student
        2. Upserts student record with curriculum_id
        3. Upserts semester records with level_semester_raw, term, level
        4. Matches each course to curriculum_course_id
        5. Upserts course records with curriculum_course_id and correct column names
        
        Args:
            student_data: Dictionary containing raw and calculated student data
        """
        supabase = self._get_supabase()
        if not supabase:
            return
        
        raw = student_data.get("raw", {})
        student_info = raw.get("student", {})
        semesters = raw.get("semesters", [])
        calculated = student_data.get("calculated", {})
        metadata = student_data.get("metadata", {})
        
        student_code = student_info.get("id")
        
        if not student_code:
            logger.warning(f"Skipping student without ID: {student_info.get('name')}")
            return
        
        try:
            enrollment_year = metadata.get("enrollment_year") or extract_enrollment_year(semesters)
            department_id = self._get_department_id()
            
            if not department_id:
                logger.warning(f"Skipping student {student_code}: No department_id found")
                return
            
            if not enrollment_year:
                logger.warning(f"Skipping student {student_code}: Could not determine enrollment_year")
                return
            
            # Resolve curriculum_id for this student
            curriculum_id = self._resolve_curriculum_with_cache(department_id, enrollment_year)
            
            if not curriculum_id:
                logger.warning(f"Skipping student {student_code}: No curriculum found for department {department_id}, enrollment {enrollment_year}")
                return
            
            now_iso = datetime.utcnow().isoformat()
            
            # Upsert student record (includes curriculum_id)
            student_result = supabase.table("students").upsert({
                "student_code": student_code,
                "name": student_info.get("name", ""),
                "department_id": department_id,
                "curriculum_id": curriculum_id,
                "enrollment_year": enrollment_year,
                "current_level": metadata.get("study_level"),
                "cumulative_gpa": calculated.get("cumulative_gpa", 0),
                "attempted_hours": calculated.get("attempted_hours", 0),
                "completed_hours": calculated.get("completed_hours", 0),
                "completion_rate": calculated.get("completion_rate", 0),
                "total_passed_courses": calculated.get("total_passed_courses", 0),
                "total_failed_courses": calculated.get("total_failed_courses", 0),
                "total_courses": calculated.get("total_courses", 0),
                "cumulative_percentage": parse_percentage(student_info.get("cumulative_percentage", "")),
                "is_active": True,
                "updated_at": now_iso
            }, on_conflict="student_code").execute()
            
            # Get student database ID
            if student_result.data:
                student_db_id = student_result.data[0]["id"]
            else:
                existing = supabase.table("students").select("id").eq("student_code", student_code).execute()
                if not existing.data:
                    logger.error(f"Could not find or create student {student_code}")
                    return
                student_db_id = existing.data[0]["id"]
            
            logger.info(f"✅ Saved/Updated student {student_code} (curriculum_id: {curriculum_id}, GPA: {calculated.get('cumulative_gpa', 0)})")
            
            # Process each semester
            for idx, semester in enumerate(semesters):
                semester_calc = semester.get("calculated", {})
                level_semester_raw = semester.get("level_semester", "")
                
                # Parse term and level from Arabic text
                term_value = parse_term_from_arabic(level_semester_raw)
                level_value = parse_level_from_arabic(level_semester_raw)
                
                # Upsert semester record with new fields
                semester_result = supabase.table("student_semesters").upsert({
                    "student_id": student_db_id,
                    "semester_number": idx + 1,
                    "academic_year": semester.get("academic_year", ""),
                    "level_semester_raw": level_semester_raw,
                    "term": term_value,
                    "level": level_value,
                    "gpa": semester_calc.get("gpa", 0),
                    "attempted_hours": semester_calc.get("attempted_hours", 0),
                    "completed_hours": semester_calc.get("completed_hours", 0),
                    "passed_courses": semester_calc.get("passed_courses", 0),
                    "failed_courses": semester_calc.get("failed_courses", 0),
                    "total_courses": semester_calc.get("total_courses", 0),
                    "quality_points": semester_calc.get("quality_points", 0),
                    "updated_at": now_iso
                }, on_conflict="student_id,semester_number").execute()
                
                # Get semester database ID
                if semester_result.data:
                    semester_db_id = semester_result.data[0]["id"]
                else:
                    existing_sem = supabase.table("student_semesters") \
                        .select("id") \
                        .eq("student_id", student_db_id) \
                        .eq("semester_number", idx + 1) \
                        .execute()
                    if not existing_sem.data:
                        continue
                    semester_db_id = existing_sem.data[0]["id"]
                
                # Process each course in the semester
                for course in semester.get("courses", []):
                    # Get grade_points from database grade_scale
                    grade_letter = course.get("grade_letter", "F")
                    grade_points = get_grade_points_from_db(supabase, grade_letter)
                    
                    # Match course to curriculum_course
                    curriculum_course_id = self._match_course_with_cache(
                        curriculum_id,
                        course.get("course_code"),
                        course.get("course_name")
                    )
                    
                    # Upsert course record with all fields
                    supabase.table("student_courses").upsert({
                        "student_id": student_db_id,
                        "semester_id": semester_db_id,
                        "curriculum_course_id": curriculum_course_id,
                        "course_code": course.get("course_code", ""),
                        "course_name": course.get("course_name", ""),
                        "credit_hours": parse_int(course.get("hours")) or 0,
                        "grade_letter": grade_letter,
                        "grade_letter_raw": course.get("grade_letter_raw", ""),
                        "grade_points": grade_points,
                        "grade_score": parse_float(course.get("score")),
                        "passed": course.get("passed", False),
                        "passed_raw": course.get("passed_raw", ""),
                        "attempt_number": course.get("attempt_number", 1),
                        "is_latest_attempt": course.get("is_latest_attempt", True),
                        "updated_at": now_iso,
                    }, on_conflict="semester_id,course_code,course_name").execute()
            
            logger.debug(f"Successfully saved/updated student {student_code} with {len(semesters)} semesters")
            
        except Exception as e:
            logger.error(f"Failed to save student {student_code}: {e}")
            self.errors.append({
                "student_code": student_code,
                "error": str(e)
            })
    
    def get_stats(self) -> Dict:
        """Get parsing statistics (students, courses, semesters counts)."""
        return self.stats