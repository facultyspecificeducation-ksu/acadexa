"""
================================================================================
Export Service for Acadexa API
================================================================================

Business logic for data export:
- Export student list to Excel
- Export at-risk students to Excel/PDF
- Export analysis issues to Excel/PDF

Security: Uses authenticated Supabase client to respect RLS.
All queries use proper joins to avoid N+1 patterns.

Author: Acadexa Team
Version: 1.1.0
================================================================================
"""

import io
import json
import logging
from datetime import datetime, timezone
from typing import Dict, List, Optional, Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import Client

from app.core.exceptions import NotFoundError

logger = logging.getLogger("acadexa.services.export")


class ExportService:
    """
    Service for data export with RLS compliance.
    
    All queries respect user permissions through authenticated client.
    Optimized with proper joins to eliminate N+1 patterns.
    """
    
    def __init__(self, supabase_client: Client):
        """
        Initialize ExportService with authenticated Supabase client.
        
        Args:
            supabase_client: Supabase client with user's JWT token.
        """
        self._supabase = supabase_client
    
    async def export_students_list(
        self,
        search: Optional[str] = None,
        department_id: Optional[UUID] = None,
        curriculum_id: Optional[UUID] = None,
        current_level: Optional[int] = None,
        academic_status: Optional[str] = None,
        risk_level: Optional[str] = None,
        enrollment_year: Optional[int] = None,
        is_active: Optional[bool] = None,
    ) -> bytes:
        """
        Export filtered student list to Excel.
        
        Generates a real XLSX file with proper formatting.
        
        Args:
            Same filters as GET /students endpoint
            
        Returns:
            Excel file bytes
        """
        # Build query with left join to latest_academic_analyses view
        query = self._supabase.table("students")\
            .select(
                "student_code, name, enrollment_year, current_level, cumulative_gpa, "
                "attempted_hours, completed_hours, completion_rate, is_active, "
                "departments(name_ar), curricula(name_ar, regulation_year), "
                "latest_academic_analyses(academic_status, risk_level, graduation_eligible)"
            )
        
        # Apply filters
        if search:
            query = query.or_(f"student_code.ilike.%{search}%,name.ilike.%{search}%")
        
        if department_id:
            query = query.eq("department_id", str(department_id))
        
        if curriculum_id:
            query = query.eq("curriculum_id", str(curriculum_id))
        
        if current_level:
            query = query.eq("current_level", current_level)
        
        if enrollment_year:
            query = query.eq("enrollment_year", enrollment_year)
        
        if is_active is not None:
            query = query.eq("is_active", is_active)
        
        if academic_status:
            query = query.eq("latest_academic_analyses.academic_status", academic_status)
        
        if risk_level:
            query = query.eq("latest_academic_analyses.risk_level", risk_level)
        
        result = query.execute()
        students = result.data or []
        
        # Flatten the nested data
        for student in students:
            analysis = student.pop("latest_academic_analyses", {})
            student["academic_status"] = analysis.get("academic_status")
            student["risk_level"] = analysis.get("risk_level")
            student["graduation_eligible"] = analysis.get("graduation_eligible")
            
            dept = student.pop("departments", {})
            student["department_name_ar"] = dept.get("name_ar")
            
            curr = student.pop("curricula", {})
            student["curriculum_name_ar"] = curr.get("name_ar")
            student["regulation_year"] = curr.get("regulation_year")
        
        # Generate Excel using openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Define headers in Arabic
        headers = [
            "رقم الطالب", "اسم الطالب", "القسم", "اللائحة", "سنة الالتحاق",
            "المستوى", "المعدل التراكمي", "الساعات المحاولة", "الساعات المكتملة",
            "نسبة الإنجاز", "الحالة الأكاديمية", "مستوى الخطر", "أهلية التخرج", "نشط"
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Status mapping for Arabic display
        status_map = {
            "good_standing": "وفق الخطة",
            "delayed": "متأخر",
            "needs_support": "يحتاج دعم",
            "probation": "إنذار أكاديمي"
        }
        risk_map = {
            "low": "منخفض",
            "medium": "متوسط",
            "high": "مرتفع"
        }
        eligible_map = {
            True: "مؤهل",
            False: "غير مؤهل"
        }
        active_map = {
            True: "نشط",
            False: "غير نشط"
        }
        
        # Write data rows
        for row_idx, student in enumerate(students, 2):
            ws.cell(row=row_idx, column=1, value=student.get("student_code", ""))
            ws.cell(row=row_idx, column=2, value=student.get("name", ""))
            ws.cell(row=row_idx, column=3, value=student.get("department_name_ar", ""))
            ws.cell(row=row_idx, column=4, value=student.get("curriculum_name_ar", ""))
            ws.cell(row=row_idx, column=5, value=student.get("enrollment_year", ""))
            ws.cell(row=row_idx, column=6, value=student.get("current_level", ""))
            ws.cell(row=row_idx, column=7, value=float(student.get("cumulative_gpa", 0)))
            ws.cell(row=row_idx, column=8, value=student.get("attempted_hours", 0))
            ws.cell(row=row_idx, column=9, value=student.get("completed_hours", 0))
            ws.cell(row=row_idx, column=10, value=float(student.get("completion_rate", 0)))
            
            status = student.get("academic_status")
            ws.cell(row=row_idx, column=11, value=status_map.get(status, status or ""))
            
            risk = student.get("risk_level")
            ws.cell(row=row_idx, column=12, value=risk_map.get(risk, risk or ""))
            
            eligible = student.get("graduation_eligible")
            ws.cell(row=row_idx, column=13, value=eligible_map.get(eligible, ""))
            
            active = student.get("is_active")
            ws.cell(row=row_idx, column=14, value=active_map.get(active, ""))
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 18
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    async def export_at_risk_students(
        self,
        department_id: Optional[UUID] = None,
        format: str = "xlsx",
    ) -> bytes:
        """
        Export at-risk students to Excel or PDF.
        
        Args:
            department_id: Optional department filter
            format: Export format (xlsx or pdf)
            
        Returns:
            File bytes
        """
        query = self._supabase.table("latest_academic_analyses")\
            .select(
                "student_id, risk_level, analyzed_at, "
                "students!inner(student_code, name, cumulative_gpa, department_id, departments!inner(name_ar))"
            )\
            .eq("risk_level", "high")
        
        if department_id:
            query = query.eq("students.department_id", str(department_id))
        
        result = query.execute()
        
        students = []
        for item in result.data or []:
            student_info = item.get("students", {})
            dept_info = student_info.get("departments", {}) if student_info else {}
            
            students.append({
                "student_code": student_info.get("student_code"),
                "name": student_info.get("name"),
                "cumulative_gpa": student_info.get("cumulative_gpa"),
                "department_name_ar": dept_info.get("name_ar"),
                "risk_level": item.get("risk_level"),
                "analyzed_at": item.get("analyzed_at"),
            })
        
        if format == "pdf":
            # For PDF, return JSON with instruction (PDF generation requires ReportLab)
            # Instead of faking, return a clear JSON response
            return json.dumps({
                "students": students,
                "total": len(students),
                "exported_at": datetime.now(timezone.utc).isoformat(),
                "format": "pdf",
                "message": "PDF generation requires ReportLab integration. Use XLSX format for Excel export."
            }).encode()
        
        # Generate XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "At-Risk Students"
        
        headers = [
            "رقم الطالب", "اسم الطالب", "القسم", "المعدل التراكمي", "مستوى الخطر", "تاريخ التحليل"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        risk_map = {"high": "مرتفع", "medium": "متوسط", "low": "منخفض"}
        
        for row_idx, student in enumerate(students, 2):
            ws.cell(row=row_idx, column=1, value=student.get("student_code", ""))
            ws.cell(row=row_idx, column=2, value=student.get("name", ""))
            ws.cell(row=row_idx, column=3, value=student.get("department_name_ar", ""))
            ws.cell(row=row_idx, column=4, value=float(student.get("cumulative_gpa", 0)))
            ws.cell(row=row_idx, column=5, value=risk_map.get(student.get("risk_level"), student.get("risk_level", "")))
            ws.cell(row=row_idx, column=6, value=student.get("analyzed_at", ""))
        
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 18
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    async def export_analysis_issues(
        self,
        student_id: UUID,
        format: str = "xlsx",
    ) -> bytes:
        """
        Export a student's analysis issues to Excel or PDF.
        
        Args:
            student_id: Student UUID
            format: Export format (xlsx or pdf)
            
        Returns:
            File bytes
        """
        # Get student info
        student_result = self._supabase.table("students")\
            .select("name, student_code")\
            .eq("id", str(student_id))\
            .execute()
        
        student_name = student_result.data[0]["name"] if student_result.data else "Unknown"
        student_code = student_result.data[0]["student_code"] if student_result.data else ""
        
        # Get latest analysis with issues
        result = self._supabase.table("latest_academic_analyses")\
            .select(
                "id, academic_status, risk_level, graduation_eligible, analyzed_at, "
                "analysis_issues(rule_code, severity, title_ar, description_ar, recommendation_ar, resolved, created_at)"
            )\
            .eq("student_id", str(student_id))\
            .execute()
        
        if not result.data:
            return json.dumps({"issues": [], "message": "No analysis found"}).encode()
        
        analysis = result.data[0]
        issues = analysis.pop("analysis_issues", []) if analysis else []
        
        if format == "pdf":
            return json.dumps({
                "student_name": student_name,
                "student_code": student_code,
                "analysis": analysis,
                "issues": issues,
                "total_issues": len(issues),
                "exported_at": datetime.now(timezone.utc).isoformat(),
                "format": "pdf",
                "message": "PDF generation requires ReportLab integration. Use XLSX format for Excel export."
            }).encode()
        
        # Generate XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Issues"
        
        # Header with student info
        ws.cell(row=1, column=1, value="تقرير المشكلات الأكاديمية")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        header_title = ws.cell(row=1, column=1)
        header_title.font = Font(bold=True, size=14)
        header_title.alignment = Alignment(horizontal="center")
        
        ws.cell(row=2, column=1, value=f"الطالب: {student_name} ({student_code})")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        
        ws.cell(row=3, column=1, value=f"تاريخ التقرير: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
        
        # Empty row
        ws.cell(row=4, column=1, value="")
        
        # Issues headers
        row_idx = 5
        headers = ["المشكلة", "الشفرة", "الشدة", "الوصف", "التوصية", "محلولة"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        severity_colors = {
            "error": "e74c3c",
            "warning": "f39c12",
            "info": "3498db"
        }
        severity_map = {
            "error": "خطأ",
            "warning": "تحذير",
            "info": "معلومة"
        }
        resolved_map = {
            True: "نعم",
            False: "لا"
        }
        
        row_idx += 1
        for issue in issues:
            severity = issue.get("severity", "info")
            color = severity_colors.get(severity, "95a5a6")
            
            ws.cell(row=row_idx, column=1, value=issue.get("title_ar", ""))
            ws.cell(row=row_idx, column=2, value=issue.get("rule_code", ""))
            
            cell = ws.cell(row=row_idx, column=3, value=severity_map.get(severity, severity))
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            ws.cell(row=row_idx, column=4, value=issue.get("description_ar", ""))
            ws.cell(row=row_idx, column=5, value=issue.get("recommendation_ar", ""))
            ws.cell(row=row_idx, column=6, value=resolved_map.get(issue.get("resolved"), "لا"))
            row_idx += 1
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20
            if col_idx == 4 or col_idx == 5:
                ws.column_dimensions[column_letter].width = 35
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()